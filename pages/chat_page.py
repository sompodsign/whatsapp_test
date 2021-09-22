import os
from selenium.webdriver.common.keys import Keys
import openpyxl as excel
import time
from utils.locators import Locators
from pages.base_page import BasePage


class ChatPage(BasePage):
    def __init__(self, driver):
        self.locator = Locators
        super(ChatPage, self).__init__(driver)  # Python2 version

    def menu_open(self):
        self.find_element(*self.locator.VERTICAL_DOT_MENU).click()
        time.sleep(1)

    def click_logout(self):
        self.find_element(*self.locator.LOGOUT).click()

    def canvas_on_page(self):
        self.wait_element(*self.locator.QR_CODE)
        canvas = self.find_element(*self.locator.QR_CODE)
        time.sleep(1)
        if canvas:
            return True
        return False

    def write_excel(self, status=None, filename=''):
        wb = excel.Workbook()
        ws = wb.active
        ws['A1'], ws['B1'] = 'Contact', 'Status'
        ws['A2'], ws['B2'] = self.read_contact('contacts.xlsx'), status
        wb.save(filename)
        wb.close()
        print(filename, "successfully created")

    def read_contact(self, filename):
        file = os.path.abspath(filename)
        file = excel.load_workbook(file)
        sheet = file.active
        firstCol = sheet['A']
        file.close()
        return firstCol[1].value

    def login(self):
        self.wait_element(*self.locator.QR_CODE)
        self.wait_element_custom_time(*self.locator.SEARCH_BOX, time=50)
        time.sleep(1)

    def search_contact(self, number):
        time.sleep(1)
        self.wait_element(*self.locator.SEARCH_BOX)
        search_field = self.find_element(*self.locator.SEARCH_BOX)
        search_field.send_keys(number)
        self.wait_element(*self.locator.MATCHED_CONTACT)
        time.sleep(1)
        return self.find_element(*self.locator.MATCHED_CONTACT)

    def send_message(self):
        contact_number = self.read_contact('contacts.xlsx')
        message = "Test Message"
        self.search_contact(contact_number).click()
        self.wait_element(*self.locator.MESSAGE_INPUT_FIELD)
        message_input_field = self.find_element(*self.locator.MESSAGE_INPUT_FIELD)
        message_input_field.send_keys(message)
        message_input_field.send_keys(Keys.ENTER)
        time.sleep(2)

    def send_message_to_multiple_numbers(self):
        contact_number = self.read_contact('contacts.xlsx')
        message = "Test Message"
        self.search_contact(contact_number).click()
        self.wait_element(*self.locator.MESSAGE_INPUT_FIELD)
        message_input_field = self.find_element(*self.locator.MESSAGE_INPUT_FIELD)
        message_input_field.send_keys(message)
        message_input_field.send_keys(Keys.ENTER)
        time.sleep(2)

    def last_message(self):
        self.wait_element(*self.locator.LAST_MESSAGE)
        message_elem = self.find_element(*self.locator.LAST_MESSAGE)
        time.sleep(2)
        if message_elem:
            return True
        return False

    def check_message_status(self):
        self.wait_element(*self.locator.LAST_MESSAGE)
        initial_status = self.find_element(*self.locator.STATUS).find_element_by_tag_name('span').get_attribute('aria-label')
        return initial_status

    def send_message_to_matched_contact(self):
        self.login()
        self.send_message()
        time.sleep(2)
        return self.last_message()

    def send_message_and_write_excel(self):
        self.login()
        self.send_message()
        if self.last_message():
            if self.check_message_status() != "Pending":
                self.write_excel(filename='message_successful.xlsx', status='Sent')

    def seen_status_to_excel(self):
        self.login()
        self.send_message()
        if self.check_message_status() == "seen":
            current_status = "seen"
        else:
            current_status = "not seen"
        self.write_excel(filename='seen_status.xlsx', status=current_status)

    def logout(self):
        self.login()
        time.sleep(1)
        self.menu_open()
        self.click_logout()
        if self.canvas_on_page():
            print("Successfully Logged out!")
        else:
            print("Something went wrong while logging out.")

